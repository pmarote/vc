import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path

def fmt_br(val) -> str:
    """Formata valores numéricos para o padrão brasileiro."""
    if val is None: return ""
    if isinstance(val, (float, int)):
        if isinstance(val, float):
            text_val = f"{val:_.2f}".replace('.', ',').replace('_', '.')
            # Adiciona cor vermelha para negativos no Markdown/HTML
            return f'<span style="color:red">{text_val}</span>' if val < 0 else text_val
        return str(val)
    return str(val).replace("\t", " ")

def export_excel(cursor: sqlite3.Cursor, out_path: Path) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    headers = [desc[0] for desc in cursor.description] if cursor.description else []
    ws.append(headers)
    
    # Estilização do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.row_dimensions[1].height = 30 
    ws.freeze_panes = "A2"
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    col_widths = [len(str(h)) + 2 for h in headers]
    row_count = 0
    
    for row in cursor:
        ws.append(row)
        row_count += 1
        current_row_idx = ws.max_row
        
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row_idx, column=col_idx)
            if isinstance(value, int) and not isinstance(value, bool):
                cell.number_format = '0' 
            elif isinstance(value, float):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
            
            if value is not None:
                curr_len = len(str(value))
                if curr_len > col_widths[col_idx-1]:
                    col_widths[col_idx-1] = curr_len

    if ws.dimensions:
        ws.auto_filter.ref = ws.dimensions
    
    for i, column_cells in enumerate(ws.columns):
        ws.column_dimensions[column_cells[0].column_letter].width = min(col_widths[i] + 2, 60)

    wb.save(out_path)
    return row_count

def export_markdown(cursor: sqlite3.Cursor, out_path: Path, sql_query: str, db_path: str = "", attachments: str = "") -> int:
    headers = [desc[0] for desc in cursor.description] if cursor.description else []
    first_row = cursor.fetchone()
    row_count = 0
    
    with open(out_path, "w", encoding="utf-8") as f:
        f.write('<details>\n  <summary><span style="font-size:0.9em; color:gray; cursor:pointer">🔍 Ver Query SQL Original</summary>\n\n')
        f.write(f"```sql\n{sql_query.strip()}\n```\n</details>\n\n")
        
        if not headers:
            f.write("> ⚠️ A consulta não retornou colunas.\n")
            return 0

        f.write("| " + " | ".join(headers) + " |\n")

        separators = []
        for i, _ in enumerate(headers):
            is_num = isinstance(first_row[i], (int, float)) if first_row else False
            separators.append("---:" if is_num else ":---")
        f.write("| " + " | ".join(separators) + " |\n")

        if first_row:
            f.write("| " + " | ".join(fmt_br(c) for c in first_row) + " |\n")
            row_count += 1

        for row in cursor:
            f.write("| " + " | ".join(fmt_br(c) for c in row) + " |\n")
            row_count += 1

        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Rodapé rico com informações dos bancos
        f.write(f"\n> 📅 **Gerado em:** {data_hora} &nbsp;|&nbsp; 🗄️ **Base:** `{db_path}`")
        if attachments:
            f.write(f" &nbsp;|&nbsp; 🔗 **Anexos:** `{attachments}`")
        f.write(f" &nbsp;|&nbsp; 📊 **Linhas:** {row_count}\n\n")

    return row_count

def export_tsv(cursor: sqlite3.Cursor, out_path: Path) -> int:
    row_count = 0
    with open(out_path, "w", encoding="utf-8") as f:
        if cursor.description:
            f.write("\t".join(d[0] for d in cursor.description) + "\n")
        
        for row in cursor:
            clean_row = [str(item).replace('.', ',') if isinstance(item, float) else str(item).replace("\t", " ") if item is not None else "" for item in row]
            f.write("\t".join(clean_row) + "\n")
            row_count += 1
            
    return row_count